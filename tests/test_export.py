from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook
from trump_monitor.collectors.sample import SampleAdapter
from trump_monitor.config import AppConfig
from trump_monitor.engine import TrumpEventEngine
from trump_monitor.exporters.excel_exporter import export_excel


def test_excel_six_sheets(tmp_path):
    root=Path(__file__).resolve().parents[1]
    result=TrumpEventEngine(AppConfig(mode="SAMPLE"),[SampleAdapter(root/"data/sample_items.json")]).run(datetime(2026,7,28,0,0,tzinfo=timezone.utc))
    out=export_excel(result,tmp_path/"report.xlsx")
    wb=load_workbook(out,read_only=True)
    assert wb.sheetnames == ["01_三日重大摘要","02_新聞明細","03_市場影響","04_產業影響","05_GTC事件輸出","06_方法與限制","07_台股候選","08_V2功能與限制","09_來源健康"]

def test_source_health_sheet_contains_nine_source_dashboard(tmp_path):
    from datetime import datetime, timezone
    from openpyxl import load_workbook
    from trump_monitor.config import AppConfig
    from trump_monitor.engine import TrumpEventEngine
    from trump_monitor.exporters.excel_exporter import export_excel
    from trump_monitor.models import RawItem
    class Feed:
        name="google_news_rss"; last_status="SUCCESS:3"; last_observations=[]
        def collect(self,start,end):
            return [
                RawItem(raw_item_id="r1",source_name="Reuters",publisher_group="Reuters",source_type="MEDIA_REPORT",published_at=end,title="Trump policy Reuters",body="Trump policy",url="https://example.com/r",source_confidence=.9,direct_quote=False,source_tier=2,source_role="VERIFICATION",acquisition_method="GOOGLE_NEWS_RSS"),
                RawItem(raw_item_id="b1",source_name="Bloomberg",publisher_group="Bloomberg",source_type="MEDIA_REPORT",published_at=end,title="Trump policy Bloomberg",body="Trump policy",url="https://example.com/b",source_confidence=.9,direct_quote=False,source_tier=2,source_role="VERIFICATION",acquisition_method="GOOGLE_NEWS_RSS"),
            ]
    result=TrumpEventEngine(AppConfig(mode="ONLINE"),[Feed()]).run(datetime.now(timezone.utc))
    path=export_excel(result,tmp_path/"out.xlsx")
    book=load_workbook(path,read_only=True,data_only=True)
    ws=book["09_來源健康"]
    labels=[ws.cell(r,1).value for r in range(4,13)]
    assert "Reuters" in labels
    assert "Bloomberg" in labels
    assert "CNBC" in labels
    assert "Truth Official" in labels
