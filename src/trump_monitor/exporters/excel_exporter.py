from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

from trump_monitor.models import RunResult

TITLE_FILL = PatternFill("solid", fgColor="17365D")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE = "FFFFFF"
BORDER = Border(*(Side(style="thin", color="D0D7DE") for _ in range(4)))


def _style_title(ws, end_col: int, text: str) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    c = ws.cell(1, 1, text)
    c.fill = TITLE_FILL
    c.font = Font(bold=True, color=WHITE, size=16)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28


def _header(ws, row: int, headers: list[str]) -> None:
    for col, value in enumerate(headers, 1):
        c = ws.cell(row, col, value)
        c.fill = HEADER_FILL
        c.font = Font(bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def _body(ws, min_row: int, max_row: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER


def _widths(ws, mapping: dict[int, float]) -> None:
    for col, width in mapping.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def export_excel(result: RunResult, path: str | Path) -> Path:
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "01_三日重大摘要"
    _style_title(ws, 21, "川普最近三天新聞與市場影響｜真正重大事件專業情報輸出")
    ws.merge_cells("A2:U2")
    ws["A2"] = f"Run ID：{result.run_id}｜Rule：{result.rule_version}｜Prompt：{result.prompt_version}｜Schema：{result.schema_version}｜Status：{result.status}｜Data Mode：{result.data_mode}｜Truth Social：{result.truth_social_status}"
    ws["A2"].fill = NOTE_FILL; ws["A2"].alignment = Alignment(wrap_text=True)
    ws.append([]); ws.append(["事件總數", len(result.events), "真正重大事件數", sum(e.is_material for e in result.events), "資料來源", ", ".join(result.source_status), "警告數", len(result.warnings)])
    headers = ["排名","事件ID","事件日期","最新更新時間","重要度","事件分數","重大性分數","重大性等級","真正重大","主題","類別","摘要重點","消息來源數","可信度","美股分數","台股分數","原油分數","黃金分數","受惠方向","受壓方向","GTC建議"]
    _header(ws, 8, headers)
    for rank, e in enumerate(result.events, 1):
        impact = {x.asset: x.final_score for x in e.impacts}
        ws.append([rank,e.event_id,e.first_seen.date().isoformat(),e.last_seen.isoformat(),"★"*e.score.importance,e.score.rule_score,e.materiality_score,e.materiality_level,"是" if e.is_material else "否",e.topic,e.category,e.summary,e.source_count,e.score.confidence,impact.get("美股",0),impact.get("台股",0),impact.get("原油",0),impact.get("黃金",0),"、".join(e.beneficiary_sectors),"、".join(e.negative_sectors),e.battle_action])
    _body(ws,9,8+len(result.events),21); _widths(ws,{1:7,2:24,3:14,4:20,5:12,6:11,7:13,8:13,9:12,10:32,11:20,12:50,13:12,14:12,15:11,16:11,17:11,18:11,19:30,20:30,21:13})
    ws.freeze_panes = "A9"

    ws2 = wb.create_sheet("02_新聞明細")
    _style_title(ws2,16,"最近三天新聞明細與來源｜證據鏈與來源優先級")
    headers2=["發布時間","事件ID","新聞標題","來源","Publisher Group","主題分類","來源可信度","來源類型","優先級Tier","驗證角色","取得方式","是否直接引用","是否重複","Duplicate Of","事件群組","來源網址"]
    _header(ws2,3,headers2)
    row=4
    for e in result.events:
        for src in e.sources:
            ws2.append([src.published_at.isoformat(),e.event_id,src.title,src.source_name,src.publisher_group,e.category,src.source_confidence,src.source_type,src.source_tier,src.source_role,src.acquisition_method,"是" if src.direct_quote else "否","否","",e.event_id,src.url]); row += 1
    _body(ws2,4,row-1,16); _widths(ws2,{1:20,2:24,3:55,4:22,5:20,6:20,7:14,8:18,9:12,10:16,11:20,12:13,13:12,14:16,15:24,16:70}); ws2.freeze_panes="A4"

    ws3 = wb.create_sheet("03_市場影響")
    _style_title(ws3,10,"市場影響推估｜Rule 70%＋AI 30%")
    _header(ws3,4,["市場／資產","Rule分數","AI分數","綜合分數","信心度","方向","主要理由","可能受惠","可能受壓","觀察期間"])
    aggregate: dict[str,list] = {}
    for e in result.events:
        for i in e.impacts: aggregate.setdefault(i.asset,[]).append(i)
    for asset, impacts in aggregate.items():
        avg=lambda field: round(sum(getattr(x,field) for x in impacts)/len(impacts))
        top=max(impacts,key=lambda x:abs(x.final_score))
        ws3.append([asset,avg("rule_score"),avg("ai_score"),avg("final_score"),sum(x.confidence for x in impacts)/len(impacts),top.direction,top.rationale,top.beneficiary,top.negative,top.horizon])
    _body(ws3,5,4+len(aggregate),10); _widths(ws3,{1:16,2:12,3:12,4:12,5:12,6:16,7:48,8:30,9:30,10:18}); ws3.freeze_panes="A5"

    ws4 = wb.create_sheet("04_產業影響")
    _style_title(ws4,11,"產業與台股傳導方向｜含信心度與GTC Gate")
    _header(ws4,3,["產業","方向","分數","信心度","資料新鮮度","驅動因素","正面催化","主要風險","GTC建議","Gate結果","備註"])
    sectors={}
    for e in result.events:
        for sec in e.beneficiary_sectors: sectors[sec]=(e,1)
        for sec in e.negative_sectors: sectors[sec]=(e,-1)
    for sec,(e,sign) in sectors.items():
        score=int(round(abs(e.score.final_score))) * sign
        direction="偏多" if score>=2 else "偏空" if score<=-2 else "中性"
        gate="BLOCKED_SAMPLE" if e.data_freshness!="CURRENT" else "WATCH_ONLY"
        ws4.append([sec,direction,score,e.score.confidence,e.data_freshness,e.category,"事件催化","事件反轉",e.battle_action,gate,"事件映射範例"])
    _body(ws4,4,3+len(sectors),11); _widths(ws4,{1:22,2:16,3:10,4:12,5:16,6:30,7:25,8:25,9:13,10:18,11:28}); ws4.freeze_panes="A4"

    ws5 = wb.create_sheet("05_GTC事件輸出")
    _style_title(ws5,22,"GTC 事件狀態機匯入格式｜gtc.trump_event.v2.2")
    hdr=["schema_version","run_id","event_id","event_date","last_seen","category","topic","importance","final_score","confidence","us_score","tw_score","oil_score","gold_score","beneficiary_sectors","negative_sectors","battle_action","event_label","source_count","source_name","source_url","data_freshness"]
    _header(ws5,3,hdr)
    for e in result.events:
        impacts={x.asset:x.final_score for x in e.impacts}; src=e.sources[0] if e.sources else None
        ws5.append([result.schema_version,result.run_id,e.event_id,e.first_seen.date().isoformat(),e.last_seen.isoformat(),e.category,e.topic,e.score.importance,e.score.final_score,e.score.confidence,impacts.get("美股",0),impacts.get("台股",0),impacts.get("原油",0),impacts.get("黃金",0),"、".join(e.beneficiary_sectors),"、".join(e.negative_sectors),e.battle_action,e.event_label,e.source_count,src.source_name if src else "",src.url if src else "",e.data_freshness])
    _body(ws5,4,3+len(result.events),22); _widths(ws5,{1:22,2:28,3:24,4:14,5:20,6:20,7:34,8:12,9:12,10:12,11:11,12:11,13:11,14:11,15:32,16:32,17:16,18:24,19:12,20:20,21:70,22:18}); ws5.freeze_panes="A4"

    ws6 = wb.create_sheet("06_方法與限制")
    _style_title(ws6,7,"報表方法、版本、判定規則與限制")
    _header(ws6,3,["項目","本報表設定","正式程式規則","風險","控制方式","驗收","備註"])
    rows=[
        ["Run Metadata",result.run_id,"每次唯一run_id","樣本誤認即時","明確狀態","run_id不可空",""],
        ["版本",f"{result.rule_version}; {result.prompt_version}; {result.model_version}; {result.schema_version}","版本必填","無法重跑","寫入所有輸出","版本齊全",""],
        ["時間窗",f"最近{result.lookback_hours}小時","UTC儲存/台北顯示","時區誤差","統一時區","邊界測試",""],
        ["來源優先順序"," → ".join(result.source_priority),"Truth第一手、主流媒體驗證、聚合補充","來源缺口","逐來源狀態與筆數","優先順序可見",""],
        ["資料來源",str(result.source_status),"多Adapter","來源失效","降級與警告","來源缺口可見",str(result.source_counts)],
        ["Truth Social",result.truth_social_status,"授權API/人工匯入/搜尋索引","未取得第一手貼文","明確狀態且不以Sample替代","狀態可見",""],
        ["市場分數","Rule 70%＋AI 30%","AI失敗Rule-only","不是價格預測","信心與期間","可拆解",""],
        ["投資限制","事件情報用途","不得單獨下單","快速反轉","人工確認","不產直接買點",""],
    ]
    for r in rows: ws6.append(r)
    _body(ws6,4,3+len(rows),7); _widths(ws6,{1:20,2:38,3:40,4:30,5:30,6:30,7:28}); ws6.freeze_panes="A4"


    ws7 = wb.create_sheet("07_台股候選")
    _style_title(ws7,7,"AI/規則台股候選｜事件產業映射；正式交易前需價格、流動性與持倉Gate")
    _header(ws7,3,["排名","代號","名稱","事件分數","建議","理由","限制"])
    for r in result.taiwan_candidates:
        ws7.append([r.get("rank"),r.get("ticker"),r.get("name"),r.get("score"),r.get("action"),r.get("reasons"),"WATCH_ONLY_NO_LIVE_MARKET_GATE"])
    _body(ws7,4,3+len(result.taiwan_candidates),7); _widths(ws7,{1:10,2:12,3:18,4:14,5:14,6:65,7:34}); ws7.freeze_panes="A4"

    ws8 = wb.create_sheet("08_V2功能與限制")
    _style_title(ws8,5,"V2功能、取得方式與限制")
    _header(ws8,3,["功能","正式實作","目前限制","降級/保護","驗收狀態"])
    for r in [
      ["Truth全文","授權API或人工匯入完整內容","搜尋索引僅snippet","content_status明確標示","PASS"],
      ["Reuters/Bloomberg摘要","摘要合法取得的RSS/API/授權內容","不繞過付費牆","無全文時只摘要snippet","PASS"],
      ["AI事件分類","Rule AI必定可用；可選外部LLM endpoint","需使用者提供AI金鑰","失敗退Rule AI","PASS"],
      ["台股受惠候選","事件→產業→股票映射與分數","未接即時行情/流動性/持倉","全部WATCH","PASS"],
      ["GTC WatchList","JSON/CSV review-required輸出","未直接寫外部GTC DB","REVIEW_REQUIRED","PASS"],
      ["Word/PDF","一鍵下載","PDF字型依部署環境","fallback字型","PASS"],
      ["每5分鐘更新","頁面開啟自動刷新＋GitHub排程","Cloud sleep/GitHub cron可能延遲","best effort","PASS"],
      ["歷史資料庫","run/event/source/impact/watchlist SQLite","Streamlit本機磁碟非永久","建議外接PostgreSQL/S3","PASS"],
    ]: ws8.append(r)
    _body(ws8,4,11,5); _widths(ws8,{1:22,2:44,3:42,4:38,5:14}); ws8.freeze_panes="A4"

    ws9 = wb.create_sheet("09_來源健康")
    _style_title(ws9,6,"來源執行狀態與筆數｜CNBC獨立可追溯")
    _header(ws9,3,["來源鍵","狀態","筆數","來源定位","是否獨立執行","備註"])
    labels={
      "truth_official_timeline":"Truth Social Official Timeline",
      "truth_official_api":"Truth Social Licensed API",
      "truth_manual_import":"Truth Social Manual Import",
      "truth_search_index":"Truth Social Search Index",
      "cnbc":"CNBC（Google News RSS source filter）",
      "google_news_rss":"Google News RSS",
      "gnews":"GNews",
      "newsapi":"NewsAPI",
    }
    all_keys=list(dict.fromkeys(list(labels)+list(result.source_status)))
    for key in all_keys:
        status=result.source_status.get(key,"NOT_CONFIGURED")
        count=result.source_counts.get(key,0)
        role="財經媒體驗證" if key=="cnbc" else "來源蒐集／補充"
        independent="是" if key in result.source_status else "否"
        warning_detail="；".join(w for w in result.warnings if w.startswith(f"{key}:"))
        note="原CNBC資料曾包含在google_news_rss內；V2.2.2新增獨立狀態與筆數。" if key=="cnbc" else ""
        if warning_detail:
            note=(note+"；" if note else "")+warning_detail
        ws9.append([key,status,count,labels.get(key,key),independent,note])
    _body(ws9,4,3+len(all_keys),6); _widths(ws9,{1:28,2:24,3:12,4:44,5:18,6:65}); ws9.freeze_panes="A4"

    if result.source_observations:
        start_row=6+len(all_keys)
        ws9.cell(start_row,1,"Truth Official 四層取得與人工查閱紀錄")
        ws9.cell(start_row,1).font=Font(bold=True,size=13,color="1F4E78")
        headers=["層級","狀態","顯示/回傳內容","備註","可進事件引擎","證據品質","官方網址","觀察時間"]
        for col,val in enumerate(headers,1): ws9.cell(start_row+1,col,val)
        _header(ws9,start_row+1,headers)
        for obs in result.source_observations:
            ws9.append([obs.layer,obs.status,obs.displayed_text,obs.note,"是" if obs.eligible_for_event_engine else "否",obs.evidence_quality,obs.url,obs.observed_at.isoformat()])
        _body(ws9,start_row+2,start_row+1+len(result.source_observations),8)
        _widths(ws9,{1:26,2:34,3:58,4:62,5:18,6:22,7:65,8:28})

    tmp = out.with_suffix(out.suffix + ".tmp")
    wb.save(tmp)
    tmp.replace(out)
    return out
